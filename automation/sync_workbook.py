"""Take a client workbook Paul has shared and update the app from it.

One command, so n8n has one thing to run:

  python automation/sync_workbook.py --file /tmp/workbook.xlsx

It works out where the client's figures sit, writes the two TypeScript modules,
checks every figure back against the workbook, and prints what it did as JSON on
stdout.

The workbook is handed in rather than fetched. Paul shares these with one Google
account and they are not public — a plain download gets a 401 — so n8n does the
fetching with that account's credentials and this script never needs any. --url
is there for working by hand on a file that happens to be open.

Nothing is committed here. n8n commits, and Netlify rebuilds — so the last word
on whether this reaches clients stays outside this script.

Two outcomes n8n needs to tell apart, and both are on stdout as {"status": …}:

  "updated"   an existing client's figures were replaced
  "new"       a client the app has never seen; the module is written and the
              registry entry is printed, but a person still has to add it

A new client is deliberately not wired in automatically. Everything else here is
checkable against the workbook; whether a business belongs to a given portal
login is not, and getting that wrong shows one client another's books.

Exit codes: 0 did something, 1 refused, 2 could not read the workbook.
"""
import argparse
import io
import json
import os
import re
import sys
import tempfile
import urllib.request

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'lib'))

import openpyxl                                                    # noqa: E402
from convert import build_modules, verify_against_workbook          # noqa: E402
from derive_rowmap import derive                                    # noqa: E402

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT_URL = 'https://docs.google.com/spreadsheets/d/{}/export?format=xlsx'
FILE_ID = re.compile(r'/spreadsheets/d/([A-Za-z0-9_-]+)')


def die(message, code=1, **extra):
    print(json.dumps({'status': 'refused', 'reason': message, **extra}))
    sys.exit(code)


def download(file_id, into):
    req = urllib.request.Request(EXPORT_URL.format(file_id),
                                 headers={'User-Agent': 'ftg-workbook-sync'})
    with urllib.request.urlopen(req, timeout=180) as r:
        data = r.read()
    if not data.startswith(b'PK'):
        # Drive hands back an HTML sign-in page when the file is not shared.
        die('That file is not readable. Share it, or set it to anyone with the link.', 2)
    with open(into, 'wb') as f:
        f.write(data)


def find_tabs(wb):
    """Which tabs hold the balance sheet and the profit and loss.

    Every client names them after themselves — "UE BS", "2G3B P&L" — so they are
    found by their ending rather than listed per client, which is the thing that
    would otherwise need editing for every new one.
    """
    bs = next((w.title for w in wb.worksheets if re.search(r'\bBS$', w.title, re.I)), None)
    pl = next((w.title for w in wb.worksheets
               if re.search(r'\b(P&L|PL|PS)$', w.title, re.I)), None)
    if not (bs and pl):
        die('Could not find the balance sheet and P&L tabs', 2,
            tabs=[w.title for w in wb.worksheets])
    return bs, pl


def client_name(wb):
    """The client's name, as their own FS-R states it."""
    ws = wb['FS-R']
    for r in range(1, 6):
        v = ws.cell(r, 2).value or ws.cell(r, 1).value
        if v and str(v).strip():
            return str(v).strip()
    return 'Client'


def slugify(name):
    s = re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')
    return re.sub(r'-(llc|inc|ltd|co)$', '', s)          # the legal suffix is noise


def camel(slug):
    head, *rest = slug.split('-')
    return head + ''.join(w.capitalize() for w in rest)


def existing_module(name, slug):
    """The module an already-registered client's figures live in.

    Their file is named after whoever set them up — Uniquely Enough's is
    ueSheets.ts — so a name derived from the workbook would not find it, and the
    automation would write a second module beside the real one and report a new
    client. The registry is asked instead: it is the list of who exists.
    """
    src = io.open(os.path.join(REPO, 'src', 'lib', 'clientDashboards.ts'),
                  encoding='utf-8').read()

    def squash(x):
        return re.sub(r'[^a-z0-9]', '', x.lower())

    want, want_slug = squash(name), squash(slug)
    for m in re.finditer(r"key:\s*'([^']+)'.*?name:\s*'([^']+)'.*?"
                         r"import\('\.\./data/(\w+)'\)", src, re.S):
        key, entry_name, module = m.group(1), m.group(2), m.group(3)
        known = squash(entry_name)
        # Either name inside the other: the registry shortens some of them
        # ("ACCESS GRANTED" for "Access Granted Education").
        if known and (known in want or want in known
                      or squash(key) in want_slug or want_slug in squash(key)):
            return key, module, entry_name
    return None, None, None


def registry_entry(slug, name, export, module, row_map):
    """What a person would paste into clientDashboards.ts for a new client."""
    rows = ',\n'.join(f'    {k}: {json.dumps(v)}' for k, v in row_map.items())
    return f'''  {{
    emails: ['SET THE CLIENT\\'S SIGN-IN EMAIL'],
    dashboard: {{
      key: '{slug}',
      label: 'Financial Dashboard',
      name: '{name.upper()}',
      subtitle: '',
      load: () => import('../data/{module}').then(m => m.{export}),
      loadNotes: () => import('../data/{module}Notes').then(m => m.{export.replace('_SHEETS', '_NOTES')}),
      rows: {{
{rows}
      }},
      forecast: 'workbook',
    }},
  }},'''


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', help='a workbook already downloaded — what n8n passes')
    ap.add_argument('--url', help='or the Google Sheets link, if it is readable')
    ap.add_argument('--file-id', help='or the Drive file id on its own')
    ap.add_argument('--dry-run', action='store_true',
                    help='report what would change without writing anything')
    ap.add_argument('--emit', action='store_true',
                    help='include the generated file contents in the JSON, for a '
                         'caller that will commit them itself')
    a = ap.parse_args()

    ours = False
    if a.file:
        path = a.file
        if not os.path.exists(path):
            die(f'No such file: {path}', 2)
    else:
        fid = a.file_id
        if not fid and a.url:
            m = FILE_ID.search(a.url)
            fid = m.group(1) if m else None
        if not fid:
            die('Send --file, or --url / --file-id if the workbook is readable')
        path = os.path.join(tempfile.gettempdir(), f'ftg-{fid}.xlsx')
        download(fid, path)
        ours = True

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if 'FS-R' not in wb.sheetnames:
            die('No FS-R tab — this does not look like a client workbook', 2,
                tabs=[w.title for w in wb.worksheets])

        bs, pl = find_tabs(wb)
        name = client_name(wb)
        slug = slugify(name)
        module = f'{camel(slug)}Sheets'
        export = re.sub(r'[^A-Z0-9]+', '_', slug.upper()) + '_SHEETS'

        row_map = derive(path)
        missing = [k for k in ('cash', 'currentAssets', 'cards', 'currentLiabilities',
                               'draws', 'equity') if row_map.get(k) is None]

        sheets_ts, notes_ts, tabs_found = build_modules(
            path, export=export, name=name, bs=bs, pl=pl)

        # Checked before anything is written. A figure that does not match the
        # workbook is not something to publish and mention afterwards.
        checks = verify_against_workbook(path, sheets_ts, notes_ts, bs, pl)
        if checks['mismatches']:
            die(f'{checks["mismatches"]} figure(s) do not match the workbook',
                1, problems=checks['problems'])

        # An existing client keeps the module they already have.
        key, known_module, known_name = existing_module(name, slug)
        if known_module:
            module = known_module.removesuffix('Sheets') + 'Sheets'
            export = None            # left as the file already declares it

        sheets_path = os.path.join(REPO, 'src', 'data', f'{module}.ts')
        notes_path = os.path.join(REPO, 'src', 'data', f'{module}Notes.ts')
        known = bool(known_module) and os.path.exists(sheets_path)

        if known:
            # Keep the export name the registry already imports, or the module
            # would compile to something nothing refers to.
            current = io.open(sheets_path, encoding='utf-8').read()
            found = re.search(r'export const (\w+): ClientSheets', current)
            if not found:
                die(f'{module}.ts has no ClientSheets export to replace', 1)
            export = found.group(1)
            sheets_ts, notes_ts, tabs_found = build_modules(
                path, export=export, name=name, bs=bs, pl=pl)
            checks = verify_against_workbook(path, sheets_ts, notes_ts, bs, pl)
            if checks['mismatches']:
                die(f'{checks["mismatches"]} figure(s) do not match the workbook',
                    1, problems=checks['problems'])

        if not a.dry_run:
            with open(sheets_path, 'w', encoding='utf-8') as f:
                f.write(sheets_ts)
            with open(notes_path, 'w', encoding='utf-8') as f:
                f.write(notes_ts)

        # Taken out of the tab report before it is published, so it reads as what
        # it is: something a person has to look at.
        off_roadmap = [w for w in [tabs_found.pop('roadmapWarning', None)] if w]

        out = {
            'status': 'updated' if known else 'new',
            'name': name, 'slug': key or slug, 'export': export,
            'matchedRegistryName': known_name,
            'files': [os.path.relpath(sheets_path, REPO).replace('\\', '/'),
                      os.path.relpath(notes_path, REPO).replace('\\', '/')],
            'tabs': {'bs': bs, 'pl': pl, **tabs_found},
            'figuresChecked': checks['figures'],
            'rowMap': row_map,
            'warnings': [f'{k} not found on the balance sheet' for k in missing] + off_roadmap,
            'dryRun': a.dry_run,
        }
        if a.emit:
            # For a caller that has no checkout to write into — n8n commits these
            # through the GitHub API rather than pushing a working copy.
            out['contents'] = {out['files'][0]: sheets_ts, out['files'][1]: notes_ts}

        if not known:
            # Wiring a new client to a login is the one thing here that cannot be
            # checked against the workbook, so it is left to a person.
            out['registryEntry'] = registry_entry(slug, name, export, module, row_map)
            out['todo'] = ('Add the entry above to src/lib/clientDashboards.ts and set '
                           "the client's sign-in email.")
        print(json.dumps(out, indent=2))
    finally:
        # Only clean up what this script downloaded; a file handed in belongs to
        # the caller.
        if ours and os.path.exists(path):
            os.unlink(path)


if __name__ == '__main__':
    main()
