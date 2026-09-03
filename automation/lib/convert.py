"""Turn a client workbook into the two TypeScript modules the app ships.

The workbook is the source of truth. Nothing here interprets the figures — it
copies out the values the workbook itself computed, keeping FS-R/FS-A row numbers
so every reference stays checkable against the spreadsheet.

Two modules, and the split is the point: one holds what the client may see, the
other holds FTG's own Findings for Review and TL;DR. The notes module is fetched
only for staff, so its text never reaches the client's browser — hiding the tab
alone left every word of it in the JavaScript they downloaded.

This is the one implementation. scripts/xlsx2ts.py is a command line over it, and
the converter service calls it too, so the app and the automation cannot drift
into producing different files from the same workbook.
"""
import datetime
import json
import re

import openpyxl

# 2025 Jan-Dec = F..Q, FY total R; 2026 Jan-Dec = T..AE, FY total AF.
C25, T25, C26, T26 = 6, 18, 20, 32

# Rows above this are the workbook's title band — the client's name, the month
# headings, the Actual/Forecast row. Chrome the dashboard draws itself.
FIRST_ROW = 7


def clean(v):
    """A cell as the TypeScript module should hold it."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%b %Y')
    if isinstance(v, str):
        return v.strip() or None
    if isinstance(v, bool):
        return v
    if isinstance(v, float):
        # Spreadsheet floats carry binary noise; the workbook displays cents.
        r = round(v, 6)
        return int(r) if r == int(r) else r
    return v


def present(wb, tab):
    """Whether the workbook still shows this tab.

    A hidden tab is one the workbook has withdrawn. Paul hid Findings for Review
    in two clients' v3 — "wala nang findings for review" — and carrying it across
    anyway would leave this app the one place it still shows.
    """
    return tab in wb.sheetnames and wb[tab].sheet_state == 'visible'


def _squash(x):
    return re.sub(r'[^a-z0-9]', '', x.lower())


def _client_name(wb):
    """The client's name, as their own FS-R states it."""
    ws = wb['FS-R']
    for r in range(1, 6):
        v = ws.cell(r, 2).value or ws.cell(r, 1).value
        if v and str(v).strip():
            return str(v).strip()
    return ''


def _client_key(name):
    """The leading words of a client's name, to match a heading against.

    The two are not written the same: FS-R calls them "Uniquely Enough
    Behavioral Health LLC" and their roadmap says "UNIQUELY ENOUGH". The front
    of the name is what holds steady, so that is what is compared.
    """
    words = [w for w in re.sub(r'[^A-Za-z0-9 ]', ' ', name).split()
             if w.lower() not in ('llc', 'inc', 'ltd', 'co')]
    return _squash(''.join(words[:2]))


def roadmap_tab(wb):
    """The roadmap that belongs to this client, and a warning if none does.

    Picking it by the tab's name is not safe. Battle Protection's v7 carries
    two: "Financial Roadmap", which is Access Granted Education's copied across
    cell for cell, and "Financial Roadmap-BPA", which is theirs. Going by the
    name would have put another client's plan on their dashboard.

    So the heading inside is read, and it has to name them. Where none does the
    roadmap is left out and said so — a client seeing no roadmap is a thing to
    fix, a client seeing another client's is not.
    """
    key = _client_key(_client_name(wb))
    found = []
    for ws in wb.worksheets:
        if ws.sheet_state != 'visible':
            continue
        if not ws.title.lower().startswith('financial roadmap'):
            continue
        head = next((str(c.value).strip() for row in ws.iter_rows(max_row=3)
                     for c in row if c.value), '')
        if key and key in _squash(head):
            return ws.title, None
        found.append(f'{ws.title} — {head[:70]}')
    if not found:
        return None, None
    return None, ('left the Financial Roadmap out: no tab heading names this client. '
                  'Found ' + '; '.join(found))


def grid(wb, tab):
    """A presentation tab, trimmed of the blank rows and columns around it."""
    ws = wb[tab]
    rows = [[clean(c.value) for c in row] for row in ws.iter_rows()]
    while rows and not any(x is not None for x in rows[-1]):
        rows.pop()
    width = max((max((i + 1 for i, x in enumerate(r) if x is not None), default=0)
                 for r in rows), default=0)
    return [r[:width] + [None] * (width - len(r)) for r in rows]


def statement(wb, tab):
    """FS-R / FS-A, keyed by the row number the workbook's formulas address."""
    ws = wb[tab]
    out = []
    for r in range(FIRST_ROW, ws.max_row + 1):
        label = next((clean(ws.cell(r, c).value) for c in range(1, 6)
                      if clean(ws.cell(r, c).value) is not None), None)
        y25 = [clean(ws.cell(r, C25 + i).value) for i in range(12)]
        y26 = [clean(ws.cell(r, C26 + i).value) for i in range(12)]
        t25, t26 = clean(ws.cell(r, T25).value), clean(ws.cell(r, T26).value)
        if label is None and not any(v is not None for v in y25 + y26 + [t25, t26]):
            continue
        out.append({'r': r, 'label': label if isinstance(label, str) else str(label or ''),
                    'y2025': y25, 't2025': t25, 'y2026': y26, 't2026': t26})
    return out


def _j(v):
    return json.dumps(v, ensure_ascii=False)


def _fmt_grid(rows):
    return '[\n' + '\n'.join('  ' + _j(r) + ',' for r in rows) + '\n]'


def _fmt_stmt(rows):
    return '[\n' + '\n'.join(
        f"  {{ r: {x['r']}, label: {_j(x['label'])},\n"
        f"    y2025: {_j(x['y2025'])}, t2025: {_j(x['t2025'])},\n"
        f"    y2026: {_j(x['y2026'])}, t2026: {_j(x['t2026'])} }},"
        for x in rows) + '\n]'


def _module(header, export, type_name, parts):
    body = f'\nexport const {export}: {type_name} = {{\n' + '\n'.join(
        f"  {k}: {v.replace(chr(10), chr(10) + '  ')}," for k, v in parts
    ) + f'\n}};\n\nexport default {export};\n'
    return header + body


def build_modules(path, export, name, bs, pl):
    """The two modules, as text, plus a note of which tabs were used.

    Returns (sheets_ts, notes_ts, report).
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    sheet_parts = [
        ('ASSUMPTIONS', _fmt_grid(grid(wb, 'ASSUMPTIONS'))),
        ("'Balance Sheet'", _fmt_grid(grid(wb, bs))),
        ("'Profit and Loss'", _fmt_grid(grid(wb, pl))),
        ("'FS-R'", _fmt_stmt(statement(wb, 'FS-R'))),
        ("'FS-A'", _fmt_stmt(statement(wb, 'FS-A'))),
    ]
    report = {'roadmap': False, 'findings': False, 'tldr': False}

    # Only where the workbook has one, and only while it is on show.
    roadmap, roadmap_warning = roadmap_tab(wb)
    if roadmap:
        sheet_parts.insert(1, ("'Financial Roadmap'", _fmt_grid(grid(wb, roadmap))))
        report['roadmap'] = True
    if roadmap_warning:
        report['roadmapWarning'] = roadmap_warning

    head = (
        "import type { ClientSheets } from './clientSheets';\n\n"
        f'// {name} — every tab of their workbook, values as the workbook itself\n'
        '// computed them. Generated from the source xlsx; edit the workbook and\n'
        '// regenerate rather than editing figures here.\n'
        '//\n'
        '// FS-R and FS-A keep their ORIGINAL row numbers. The workbook\'s formulas\n'
        '// address rows, and each client numbers them differently, so the row map\n'
        '// lives with the client\'s entry in clientDashboards.ts.\n'
    )
    sheets_ts = _module(head, export, 'ClientSheets', sheet_parts)

    notes_parts = []
    for tab, key in (('Findings for Review', 'findings'), ('TL;DR', 'tldr')):
        if present(wb, tab):
            notes_parts.append((f"'{tab}'", _fmt_grid(grid(wb, tab))))
            report[key] = True

    notes_head = (
        "import type { ClientNotes } from './clientSheets';\n\n"
        f"// {name} — FTG's working notes.\n"
        '//\n'
        '// Kept apart from the statements deliberately. These are drafts about the\n'
        "// client's own bookkeeping, and this module is imported only when a staff\n"
        '// member is viewing, so none of it reaches the client at all.\n'
    )
    notes_ts = _module(notes_head, export.replace('_SHEETS', '_NOTES'),
                       'ClientNotes', notes_parts)

    return sheets_ts, notes_ts, report


# ── Checking what was generated against the workbook it came from ─────────────

def _norm(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%b %Y')
    if isinstance(v, str):
        return v.strip() or None
    if isinstance(v, float):
        r = round(v, 6)
        return int(r) if r == int(r) else r
    return v


def _same(a, b):
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) and not isinstance(a, bool):
        return abs(a - b) < 5e-7
    return a == b


# A quoted run in the generated module: a double-quoted value, or a
# single-quoted key. Matched so the rewriting below can step over them.
_QUOTED = re.compile(r'"(?:[^"\\]|\\.)*"' r"|'(?:[^'\\]|\\.)*'")


def _values(ts):
    """The data out of a generated module, without running JavaScript.

    The text is JSON but for its unquoted keys and trailing commas, so those two
    are rewritten — but only outside the quoted runs. A client's own words can
    look exactly like the things being rewritten: 2G3B name their roadmap
    "2G3B EATS, LLC: FP&A AND CONTROLLER FINANCIAL ROADMAP", and a rewrite that
    did not step over strings read ", LLC:" as a key and mangled the line.
    """
    body = ts[ts.index('= {'):]
    body = body[body.index('{'):body.rindex('}') + 1]

    def unwrapped(code):
        code = re.sub(r'(\{|,)(\s*)([A-Za-z_]\w*)\s*:', r'\1\2"\3":', code)
        return re.sub(r',(\s*[}\]])', r'\1', code)

    out, at = [], 0
    for m in _QUOTED.finditer(body):
        out.append(unwrapped(body[at:m.start()]))
        run = m.group(0)
        # A single-quoted key becomes the JSON string it stands for.
        out.append(json.dumps(run[1:-1].replace("\\'", "'"))
                   if run.startswith("'") else run)
        at = m.end()
    out.append(unwrapped(body[at:]))
    return json.loads(''.join(out))


def verify_against_workbook(path, sheets_ts, notes_ts, bs, pl):
    """Every figure in the generated modules, checked back against the workbook.

    Run before anything is published. A client's figures reaching the app
    unchecked is the failure worth ruling out, and it is cheap to rule out.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    data = _values(sheets_ts) | _values(notes_ts)

    checked = mismatches = 0
    problems = []

    for tab in ('FS-R', 'FS-A'):
        ws = wb[tab]
        for row in data[tab]:
            r = row['r']
            for field, col in (('y2025', C25), ('y2026', C26)):
                for i in range(12):
                    want, got = _norm(ws.cell(r, col + i).value), row[field][i]
                    checked += 1
                    if not _same(want, got):
                        mismatches += 1
                        problems.append(f'{tab} r{r} {field}[{i}]: {want!r} vs {got!r}')
            for field, col in (('t2025', T25), ('t2026', T26)):
                want, got = _norm(ws.cell(r, col).value), row[field]
                checked += 1
                if not _same(want, got):
                    mismatches += 1
                    problems.append(f'{tab} r{r} {field}: {want!r} vs {got!r}')

    tabs = [('Findings for Review', 'Findings for Review'), ('TL;DR', 'TL;DR'),
            ('ASSUMPTIONS', 'ASSUMPTIONS'),
            ('Financial Roadmap', roadmap_tab(wb)[0] or 'Financial Roadmap'),
            ('Balance Sheet', bs), ('Profit and Loss', pl)]
    for key, tab in [(k, t) for k, t in tabs if k in data]:
        ws = wb[tab]
        for ri, row in enumerate(data[key], start=1):
            for ci, got in enumerate(row, start=1):
                want = _norm(ws.cell(ri, ci).value)
                checked += 1
                if not _same(want, got):
                    mismatches += 1
                    problems.append(f'{key} [{ri},{ci}]: {want!r} vs {got!r}')

    return {'figures': checked, 'mismatches': mismatches, 'problems': problems[:20]}
