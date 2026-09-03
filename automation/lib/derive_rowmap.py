"""Work out where a client's figures sit in their FS-R, by reading the labels.

Until now each client's row map was read off their workbook by hand and typed
into clientDashboards.ts. That is what stopped a new client being added without
someone in the loop, and it is the thing most likely to go quietly wrong: FTG's
template is not numbered the same for everyone, and a map that is one row out
does not fail — it reports the neighbouring line.

Rows are found by what they are called instead. The labels are stable across the
workbooks in a way the numbers are not: every one of them says TOTAL REVENUE and
NET INCOME, whatever row those land on.

Run directly to check it against the six maps that were read by hand:

  python scripts/derive_rowmap.py
"""
import os
import re
import sys

import openpyxl

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Anything in these first columns can hold a row's name; the workbooks indent by
# moving the label right rather than by padding it.
LABEL_COLS = range(1, 6)

PAYROLLISH = re.compile(
    r'payroll|contract labor|casual labor'
    r"|employee benefits|workers.? comp|officer'?s salary",
    re.I)


def labels(ws, first=7):
    """Row number → its name, for every row that has one."""
    out = {}
    for r in range(first, ws.max_row + 1):
        for c in LABEL_COLS:
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                out[r] = str(v).strip()
                break
    return out


def find(rows, *patterns, after=0, before=10 ** 6):
    """The first row in range whose name matches, or None."""
    for r in sorted(rows):
        if not (after < r < before):
            continue
        for p in patterns:
            if re.fullmatch(p, rows[r], re.I):
                return r
    return None


def named_between(rows, start, end):
    """The named rows strictly between two markers."""
    return [r for r in sorted(rows) if start < r < end]


def span_between(start, end):
    """Every row between two markers, named or not.

    What a section's total actually sums over. 1st Step's other expenses run to a
    blank line its total still covers; leaving it out would make the detail add
    up to less than the total it belongs to.
    """
    return list(range(start + 1, end))


def derive(path, fs_r='FS-R'):
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = labels(wb[fs_r])

    total_income = find(rows, r'TOTAL REVENUE', r'TOTAL INCOME')
    total_opex = find(rows, r'TOTAL EXPENSES?', r'TOTAL OPERATING EXPENSES?')
    net_income = find(rows, r'NET INCOME')
    if not (total_income and total_opex and net_income):
        raise ValueError(f'{os.path.basename(path)}: no TOTAL REVENUE / TOTAL EXPENSES / NET INCOME')

    revenue_head = find(rows, r'REVENUE', r'INCOME', before=total_income)
    income = named_between(rows, revenue_head, total_income)

    total_cos = find(rows, r'TOTAL COST OF (SERVICES|GOODS SOLD)', before=total_opex)

    expenses_head = find(rows, r'EXPENSES?', after=(total_cos or total_income), before=total_opex)
    opex_rows = named_between(rows, expenses_head, total_opex)
    opex_first, opex_last = opex_rows[0], opex_rows[-1]

    # A subtotal sitting among the line items — STEER's "Total Payroll Block".
    # Counted as spending it would double what it summarises.
    opex_skip = [r for r in opex_rows if re.match(r'total\b', rows[r], re.I)]

    # The payroll block: the longest run of adjacent labour lines. It is a run
    # rather than a filter because "Bonus" and "Employee Meals" sit just below it
    # in one workbook and are not part of what the forecast pools.
    runs, run = [], []
    for r in opex_rows:
        if r in opex_skip:
            if run:
                runs.append(run); run = []
            continue
        if PAYROLLISH.search(rows[r]) and (not run or r == run[-1] + 1):
            run.append(r)
        elif PAYROLLISH.search(rows[r]):
            runs.append(run); run = [r]
        elif run:
            runs.append(run); run = []
    if run:
        runs.append(run)
    payroll = max(runs, key=len) if runs else []

    gross_profit = find(rows, r'NET OPERATING INCOME', after=total_opex, before=net_income)
    total_other_income = find(rows, r'TOTAL OTHER INCOME', after=total_opex, before=net_income)
    total_other_expense = find(rows, r'TOTAL OTHER EXPENSES?', after=total_opex, before=net_income)
    net_other = find(rows, r'NET OTHER INCOME', after=total_opex, before=net_income)

    other_income, other_expense = None, None
    if total_other_income:
        head = find(rows, r'OTHER INCOME', after=total_opex, before=total_other_income)
        other_income = span_between(head, total_other_income)
    if total_other_expense:
        head = find(rows, r'OTHER EXPENSES?', after=(total_other_income or total_opex),
                    before=total_other_expense)
        other_expense = span_between(head, total_other_expense)

    # The balance sheet, below the income statement.
    bs = find(rows, r'BALANCE SHEET', after=net_income) or net_income
    cash = find(rows, r'BANK ACCOUNTS', after=bs)
    current_assets = find(rows, r'TOTAL CURRENT ASSETS', after=bs)
    cards = find(rows, r'CREDIT CARDS', after=bs)
    current_liabs = find(rows, r'TOTAL CURRENT LIABILITIES', after=bs)
    equity = find(rows, r'TOTAL EQUITY', after=bs)
    draws = find(rows, r'.*(DRAW|PERSONAL EXPENSE|DISTRIBUTION).*', after=bs, before=equity or 10 ** 6)

    out = {
        'income': income, 'totalIncome': total_income,
        'headlineExpense': [x for x in (total_cos, total_opex, total_other_expense) if x],
        'payroll': payroll,
        'opexFirst': opex_first, 'opexLast': opex_last, 'totalOpex': total_opex,
        'netIncome': net_income,
        'cash': cash, 'currentAssets': current_assets, 'cards': cards,
        'currentLiabilities': current_liabs, 'draws': draws, 'equity': equity,
    }
    if opex_skip:
        out['opexSkip'] = opex_skip
    for k, v in (('grossProfit', gross_profit), ('otherIncome', other_income),
                 ('totalOtherIncome', total_other_income), ('otherExpense', other_expense),
                 ('totalOtherExpense', total_other_expense), ('netOther', net_other)):
        if v:
            out[k] = v
    return out


if __name__ == '__main__':
    W = 'C:/Users/Ongoing/Desktop/ftg-workbooks'
    for f in ['ue-v3.xlsx', '1sg.xlsx', '2g3b.xlsx', 'age-v3.xlsx', 'bpa-v6.xlsx', 'steer.xlsx']:
        try:
            m = derive(os.path.join(W, f))
            print(f'{f}:')
            for k, v in m.items():
                print(f'    {k:20} {v}')
        except Exception as e:
            print(f'{f}: FAILED — {e}')
        print()
