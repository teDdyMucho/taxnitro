"""Every figure in a client's generated modules, checked back against its workbook.

Reads the .ts as data rather than as text, so this compares numbers instead of
formatting. Run it after xlsx2ts.py, and again whenever a workbook is reshared —
a client's figures reaching the app unchecked is the failure worth ruling out.

Both modules are checked: the statements the client sees, and the notes module
beside it (<module>Notes.ts) that only staff are sent.

  python scripts/verify-client-sheets.py <workbook.xlsx> <module.ts> <BS tab> <PL tab>

Prints the number of figures compared and every disagreement.
"""
import json, os, re, subprocess, sys, datetime
import openpyxl
sys.stdout.reconfigure(encoding='utf-8')

xlsx, ts, bs_tab, pl_tab = sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4]


def load(path):
    """A generated module's values, read through node so its syntax decides."""
    src = re.sub(r'^import type.*$', '', open(path, encoding='utf-8').read(), flags=re.M)
    src = src.replace(': ClientSheets', '').replace(': ClientNotes', '')
    tmp = os.path.abspath(path) + '.check.mjs'
    open(tmp, 'w', encoding='utf-8').write(src)
    url = 'file:///' + tmp.replace(chr(92), '/')
    proc = subprocess.run(
        ['node', '--input-type=module', '-e',
         'const m = await import("' + url + '"); console.log(JSON.stringify(m.default));'],
        capture_output=True, text=True, encoding='utf-8')
    os.remove(tmp)
    if proc.returncode:
        print(proc.stderr[:1500]); sys.exit(1)
    return json.loads(proc.stdout)


data = load(ts)
notes_path = (ts[:-3] if ts.endswith('.ts') else ts) + 'Notes.ts'
if not os.path.exists(notes_path):
    print(f'missing notes module: {notes_path}'); sys.exit(1)
data.update(load(notes_path))

wb = openpyxl.load_workbook(xlsx, data_only=True)
C25, T25, C26, T26 = 6, 18, 20, 32


def norm(v):
    if v is None: return None
    if isinstance(v, (datetime.datetime, datetime.date)): return v.strftime('%b %Y')
    if isinstance(v, str): return v.strip() or None
    if isinstance(v, float):
        r = round(v, 6)
        return int(r) if r == int(r) else r
    return v


def same(a, b):
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) and not isinstance(a, bool):
        return abs(a - b) < 5e-7
    return a == b


checked = mismatch = 0
problems = []

for tab, key in [('FS-R', 'FS-R'), ('FS-A', 'FS-A')]:
    ws = wb[tab]
    for row in data[key]:
        r = row['r']
        for field, col in [('y2025', C25), ('y2026', C26)]:
            for i in range(12):
                want, got = norm(ws.cell(r, col + i).value), row[field][i]
                checked += 1
                if not same(want, got):
                    mismatch += 1
                    problems.append(f'{tab} r{r} {field}[{i}]: workbook {want!r} vs module {got!r}')
        for field, col in [('t2025', T25), ('t2026', T26)]:
            want, got = norm(ws.cell(r, col).value), row[field]
            checked += 1
            if not same(want, got):
                mismatch += 1
                problems.append(f'{tab} r{r} {field}: workbook {want!r} vs module {got!r}')

# Only the tabs this client's modules actually carry — a workbook can add one
# (Financial Roadmap) or withdraw one (Findings for Review, hidden in v3).
TABS = [('Findings for Review', 'Findings for Review'), ('TL;DR', 'TL;DR'),
        ('ASSUMPTIONS', 'ASSUMPTIONS'), ('Financial Roadmap', 'Financial Roadmap'),
        ('Balance Sheet', bs_tab), ('Profit and Loss', pl_tab)]
for key, tab in [(k, t) for k, t in TABS if k in data]:
    ws = wb[tab]
    for ri, row in enumerate(data[key], start=1):
        for ci, got in enumerate(row, start=1):
            want = norm(ws.cell(ri, ci).value)
            checked += 1
            if not same(want, got):
                mismatch += 1
                problems.append(f'{key} [{ri},{ci}]: workbook {want!r} vs module {got!r}')

print(f'checked {checked} figures against the workbook — {mismatch} mismatch(es)')
for p in problems[:20]:
    print('  ' + p)
