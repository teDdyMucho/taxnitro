"""Turn one FTG client workbook into a src/data/<client>Sheets.ts module.

The workbook is the source of truth. Nothing here interprets the figures — it
copies out the values the workbook itself computed, keeping FS-R/FS-A row
numbers so every reference stays checkable against the spreadsheet.

  python xlsx2ts.py <workbook.xlsx> <EXPORT_NAME> <out.ts> [--bs TAB] [--pl TAB]
"""
import sys, json, datetime, argparse
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

ap = argparse.ArgumentParser()
ap.add_argument('xlsx'); ap.add_argument('export'); ap.add_argument('out')
ap.add_argument('--bs', required=True, help='balance-sheet tab name')
ap.add_argument('--pl', required=True, help='profit-and-loss tab name')
ap.add_argument('--name', default='', help='client name for the header comment')
ap.add_argument('--first-row', type=int, default=7,
                help='first FS-R/FS-A row of real content')
a = ap.parse_args()

wb = openpyxl.load_workbook(a.xlsx, data_only=True)

# 2025 Jan-Dec = F..Q, FY total R; 2026 Jan-Dec = T..AE, FY total AF.
C25, T25, C26, T26 = 6, 18, 20, 32


def clean(v):
    """A cell as the TypeScript module should hold it."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%b %Y')
    if isinstance(v, str):
        s = v.strip()
        return s or None
    if isinstance(v, bool):
        return v
    if isinstance(v, float):
        # Spreadsheet floats carry binary noise; the workbook displays cents.
        r = round(v, 6)
        return int(r) if r == int(r) else r
    return v


def grid(tab):
    """A presentation tab, trimmed of the blank rows and columns around it."""
    ws = wb[tab]
    rows = [[clean(c.value) for c in row] for row in ws.iter_rows()]
    while rows and not any(x is not None for x in rows[-1]):
        rows.pop()
    width = max((max((i + 1 for i, x in enumerate(r) if x is not None), default=0)
                 for r in rows), default=0)
    return [r[:width] + [None] * (width - len(r)) for r in rows]


def statement(tab):
    """FS-R / FS-A, keyed by the row number the workbook's formulas address.

    Starts below the title band. Those top rows hold the client name and the
    month headings, which are chrome the dashboard draws itself — carried into
    the data they would render as unnamed rows among the figures.
    """
    ws = wb[tab]
    out = []
    for r in range(a.first_row, ws.max_row + 1):
        label = next((clean(ws.cell(r, c).value) for c in range(1, 6)
                      if clean(ws.cell(r, c).value) is not None), None)
        y25 = [clean(ws.cell(r, C25 + i).value) for i in range(12)]
        y26 = [clean(ws.cell(r, C26 + i).value) for i in range(12)]
        t25, t26 = clean(ws.cell(r, T25).value), clean(ws.cell(r, T26).value)
        if label is None and not any(v is not None for v in y25 + y26 + [t25, t26]):
            continue
        out.append({'r': r, 'label': label if isinstance(label, str) else str(label or ''),
                    'y2025': y25, 't2025': t25, 'y2026': y26, 't2026': t26})
    return out


def j(v):
    return json.dumps(v, ensure_ascii=False)


def fmt_grid(rows):
    return '[\n' + '\n'.join('  ' + j(r) + ',' for r in rows) + '\n]'


def fmt_stmt(rows):
    return '[\n' + '\n'.join(
        f"  {{ r: {x['r']}, label: {j(x['label'])},\n"
        f"    y2025: {j(x['y2025'])}, t2025: {j(x['t2025'])},\n"
        f"    y2026: {j(x['y2026'])}, t2026: {j(x['t2026'])} }},"
        for x in rows) + '\n]'


name = a.name or a.export
head = f"""import type {{ ClientSheets }} from './clientSheets';

// {name} — every tab of their workbook, values as the workbook itself computed
// them. Generated from the source xlsx by scripts/xlsx2ts.py; edit the workbook
// and regenerate rather than editing figures here.
//
// FS-R and FS-A keep their ORIGINAL row numbers. The workbook's formulas address
// rows, and each client numbers them differently, so the row map lives with the
// client's entry in clientDashboards.ts rather than in the model.
"""

def module(path, header, export, type_name, parts):
    body = f"\nexport const {export}: {type_name} = {{\n" + '\n'.join(
        f"  {k}: {v.replace(chr(10), chr(10) + '  ')}," for k, v in parts
    ) + f"\n}};\n\nexport default {export};\n"
    open(path, 'w', encoding='utf-8').write(header + body)
    print(f"{path}  {(len(header + body) / 1024):.0f}KB")
    for k, v in parts:
        n = v.count('\n  {') if 'r:' in v[:200] else v.count('\n  [')
        print(f"  {k:24} {n} rows")


# What the client may see: their own statements, and the assumptions behind the
# forecast drawn from them.
module(a.out, head, a.export, 'ClientSheets', [
    ('ASSUMPTIONS', fmt_grid(grid('ASSUMPTIONS'))),
    ("'Balance Sheet'", fmt_grid(grid(a.bs))),
    ("'Profit and Loss'", fmt_grid(grid(a.pl))),
    ("'FS-R'", fmt_stmt(statement('FS-R'))),
    ("'FS-A'", fmt_stmt(statement('FS-A'))),
])

# FTG's own notes, in a module of their own so that they are never part of what
# a client's browser downloads.
notes_out = (a.out[:-3] if a.out.endswith('.ts') else a.out) + 'Notes.ts'
notes_head = (
    "import type { ClientNotes } from './clientSheets';\n\n"
    f"// {name} — FTG's working notes, generated by scripts/xlsx2ts.py.\n"
    "//\n"
    "// Kept apart from the statements deliberately. These are drafts about the\n"
    "// client's own bookkeeping, and this module is imported only when a staff\n"
    "// member is viewing, so none of it reaches the client at all.\n"
)
module(notes_out, notes_head, a.export.replace('_SHEETS', '_NOTES'),
       'ClientNotes', [
    ("'Findings for Review'", fmt_grid(grid('Findings for Review'))),
    ("'TL;DR'", fmt_grid(grid('TL;DR'))),
])
